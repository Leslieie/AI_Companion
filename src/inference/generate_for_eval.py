"""Generate SFT responses for the human evaluation sheet.

Loads the LoRA-adapted model, runs the full stateful pipeline on each
of the 50 user messages from the evaluation spreadsheet, and saves:
  - outputs/eval/sft_human_eval_responses.jsonl  (structured records)
  - outputs/eval/human_evaluation_sft_updated.xlsx  (sheet with col M filled,
    cols N-Q cleared for fresh human scoring)

Usage:
    python -m src.inference.generate_for_eval
"""

import os
os.environ["PYTHONUTF8"] = "1"

import json
from copy import copy
from pathlib import Path

import openpyxl
from peft import PeftModel

from src.inference.generate import MODEL_NAME, generate_response, load_model
from src.inference.prompt_builder import build_prompt
from src.modules.emotion_classifier import classify_emotion
from src.modules.memory_store import MemoryStore
from src.modules.policy_selector import select_policy
from src.modules.state_tracker import StateTracker

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SHEET_PATH = PROJECT_ROOT / "data" / "human_evaluation_sheet.xlsx"
SFT_PATH = PROJECT_ROOT / "outputs" / "checkpoints" / "sft_run_01" / "final"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "eval"

FIRST_DATA_ROW = 3
LAST_DATA_ROW = 52
COL_B = 2
COL_M = 13
COL_N = 14
COL_Q = 17


def main() -> None:
    # ── Load user messages from xlsx ──
    wb = openpyxl.load_workbook(str(SHEET_PATH))
    ws = wb["Evaluation"]

    user_messages: list[tuple[int, str]] = []
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        val = ws.cell(row=row, column=COL_B).value
        if val:
            user_messages.append((row, str(val).strip()))
    print(f"Loaded {len(user_messages)} user messages from sheet\n")

    # ── Load SFT model ──
    print("Loading base model + LoRA adapter ...")
    tokenizer, base_model = load_model(MODEL_NAME)
    model = PeftModel.from_pretrained(base_model, str(SFT_PATH))
    print("Model ready.\n")

    # ── Generate responses ──
    records: list[dict] = []
    responses_by_row: dict[int, str] = {}

    for i, (row, user_msg) in enumerate(user_messages):
        tracker = StateTracker()
        memory = MemoryStore()
        state = tracker.get_state()

        state = tracker.update(user_msg)
        memory.extract_and_store(user_msg, turn=0)

        emotion = classify_emotion(user_msg)
        policy = select_policy(emotion, state)
        memories = memory.retrieve(user_msg, top_k=3)

        history = [{"role": "user", "content": user_msg}]
        prompt_messages = build_prompt(history, state, memories, policy)

        response = generate_response(tokenizer, model, prompt_messages)

        records.append({
            "id": i + 1,
            "user_message": user_msg,
            "sft_response": response,
            "emotion": emotion,
            "policy": policy,
        })
        responses_by_row[row] = response

        print(f"[{i + 1}/{len(user_messages)}] User: \"{user_msg}\"")
        print(f"  SFT: \"{response}\"")
        print("  ---")

    # ── Save JSONL ──
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    jsonl_path = OUTPUT_DIR / "sft_human_eval_responses.jsonl"
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"\nSaved {len(records)} records to {jsonl_path}")

    # ── Update xlsx: clear M-Q for all rows, fill M with new responses ──
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for col in range(COL_M, COL_Q + 1):
            ws.cell(row=row, column=col).value = None

    for row, response in responses_by_row.items():
        ws.cell(row=row, column=COL_M).value = response

    xlsx_out = OUTPUT_DIR / "human_evaluation_sft_updated.xlsx"
    wb.save(str(xlsx_out))
    print(f"Saved updated sheet to {xlsx_out}")
    print(f"  Column M: filled with {len(responses_by_row)} SFT responses")
    print(f"  Columns N-Q: cleared for fresh human scoring")


if __name__ == "__main__":
    main()
