"""Generate 7B responses (plain, stateful, stateful+SFT) for human evaluation.

Creates outputs/eval/human_evaluation_7b.xlsx and
outputs/eval/7b_human_eval_responses.jsonl.

Usage:
    python -m src.inference.generate_for_eval_7b
"""

import os
os.environ["PYTHONUTF8"] = "1"

import gc
import json
from pathlib import Path

import openpyxl
import torch
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from peft import PeftModel
from transformers import AutoModelForCausalLM, AutoTokenizer

from src.inference.generate import generate_response
from src.inference.prompt_builder import build_prompt
from src.modules.emotion_classifier import classify_emotion
from src.modules.memory_store import MemoryStore
from src.modules.policy_selector import select_policy
from src.modules.state_tracker import StateTracker

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SHEET_PATH = PROJECT_ROOT / "data" / "human_evaluation_sheet.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "eval"
SFT_7B_PATH = PROJECT_ROOT / "outputs" / "checkpoints" / "sft_7b_run_01" / "final"

MODEL_7B = "Qwen/Qwen2.5-7B-Instruct"
PLAIN_SYSTEM_PROMPT = "You are a warm AI companion."

FIRST_ROW = 3
LAST_ROW = 52


def _load_user_messages() -> list[str]:
    wb = openpyxl.load_workbook(str(SHEET_PATH), read_only=True)
    ws = wb["Evaluation"]
    msgs: list[str] = []
    for row in range(FIRST_ROW, LAST_ROW + 1):
        val = ws.cell(row=row, column=2).value
        msgs.append(str(val).strip() if val else "")
    wb.close()
    return msgs


def _load_model(adapter_path: str | None = None):
    tokenizer = AutoTokenizer.from_pretrained(MODEL_7B)
    model = AutoModelForCausalLM.from_pretrained(
        MODEL_7B, torch_dtype="auto", device_map="auto",
    )
    if adapter_path:
        model = PeftModel.from_pretrained(model, adapter_path)
    return tokenizer, model


def _free_model(model, tokenizer):
    del model, tokenizer
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()


def _generate_plain(tokenizer, model, user_msgs: list[str]) -> list[dict]:
    records: list[dict] = []
    for i, msg in enumerate(user_msgs):
        prompt = [
            {"role": "system", "content": PLAIN_SYSTEM_PROMPT},
            {"role": "user", "content": msg},
        ]
        resp = generate_response(tokenizer, model, prompt)
        records.append({
            "id": i + 1, "mode": "plain",
            "user_message": msg, "response": resp,
            "emotion": "n/a", "policy": "n/a",
        })
        print(f"  [plain {i+1}/{len(user_msgs)}] done")
    return records


def _generate_stateful(tokenizer, model, user_msgs: list[str], mode_label: str) -> list[dict]:
    records: list[dict] = []
    for i, msg in enumerate(user_msgs):
        tracker = StateTracker()
        memory = MemoryStore()
        state = tracker.update(msg)
        memory.extract_and_store(msg, turn=0)

        emotion = classify_emotion(msg)
        policy = select_policy(emotion, state)
        memories = memory.retrieve(msg, top_k=3)
        history = [{"role": "user", "content": msg}]
        prompt = build_prompt(history, state, memories, policy)

        resp = generate_response(tokenizer, model, prompt)
        records.append({
            "id": i + 1, "mode": mode_label,
            "user_message": msg, "response": resp,
            "emotion": emotion, "policy": policy,
        })
        print(f"  [{mode_label} {i+1}/{len(user_msgs)}] done")
    return records


def _build_xlsx(
    user_msgs: list[str],
    plain: list[dict],
    stateful: list[dict],
    sft: list[dict],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()

    # ── Instructions sheet ──
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    instructions = [
        ("A1", "Human Evaluation Guide — AI Companion 7B Model"),
        ("A3", "Purpose"),
        ("B3", "Rate responses from 3 model variants (7B) on the same 50 test inputs to measure improvement."),
        ("A5", "Variants"),
        ("A6", "  A"), ("B6", "Plain baseline — Qwen2.5-7B with a simple system prompt, no state/memory/policy"),
        ("A7", "  B"), ("B7", "Stateful — same 7B model + state tracker + memory store + policy selector + prompt builder"),
        ("A8", "  C"), ("B8", "Stateful + SFT — same as B but with LoRA fine-tuned 7B model"),
        ("A10", "Scoring (1–5 scale)"),
        ("A11", "  1 = Very Poor"), ("B11", "Completely off, robotic, or inappropriate"),
        ("A12", "  2 = Poor"), ("B12", "Somewhat relevant but clearly lacking warmth/coherence"),
        ("A13", "  3 = Acceptable"), ("B13", "Reasonable response, nothing notably good or bad"),
        ("A14", "  4 = Good"), ("B14", "Warm, coherent, appropriate — minor issues at most"),
        ("A15", "  5 = Excellent"), ("B15", "Natural, empathetic, persona-consistent, would continue chatting"),
        ("A17", "Dimensions"),
        ("A18", "  Warmth"), ("B18", "Does the response feel caring and emotionally supportive?"),
        ("A19", "  Coherence"), ("B19", "Does the response make sense given the user input?"),
        ("A20", "  Persona"), ("B20", "Does it sound like Ari (warm, gentle, slightly playful), not a customer service bot?"),
        ("A21", "  Engagement"), ("B21", "Would you want to keep talking after this response?"),
        ("A22", "  Emotion Match"), ("B22", "Is the tone appropriate for the user's emotional state?"),
        ("A24", "Rules"),
        ("A25", "  1."), ("B25", "Responses are shuffled — you do NOT know which variant you are rating."),
        ("A26", "  2."), ("B26", "Rate each response independently. Do not compare across columns."),
        ("A27", "  3."), ("B27", "Read the user message carefully before scoring."),
        ("A28", "  4."), ("B28", "Use the full 1–5 range. Avoid rating everything 3."),
        ("A29", "  5."), ("B29", "Add optional comments in the Notes column for anything notable."),
    ]
    for cell_ref, value in instructions:
        ws_inst[cell_ref] = value
    ws_inst["A1"].font = Font(bold=True, size=14)

    # ── Evaluation sheet ──
    ws_eval = wb.create_sheet("Evaluation")

    # Row 1: merged headers
    header_bold = Font(bold=True, size=11)
    center = Alignment(horizontal="center")

    ws_eval.merge_cells("A1:B1")
    ws_eval["A1"] = "Test Input"
    ws_eval["A1"].font = header_bold
    ws_eval["A1"].alignment = center

    for start_col, label in [(3, "Variant A (Plain)"), (8, "Variant B (Stateful)"), (13, "Variant C (Stateful+SFT)")]:
        end_col = start_col + 4
        ws_eval.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col,
        )
        cell = ws_eval.cell(row=1, column=start_col, value=label)
        cell.font = header_bold
        cell.alignment = center

    ws_eval.cell(row=1, column=18, value="Notes").font = header_bold

    # Row 2: sub-headers
    sub_headers = [
        "ID", "User Message",
        "Response", "Warmth", "Coherence", "Persona", "Engage",
        "Response", "Warmth", "Coherence", "Persona", "Engage",
        "Response", "Warmth", "Coherence", "Persona", "Engage",
        "Notes",
    ]
    sub_bold = Font(bold=True, size=10)
    for col_idx, name in enumerate(sub_headers, start=1):
        cell = ws_eval.cell(row=2, column=col_idx, value=name)
        cell.font = sub_bold

    # Rows 3-52: data
    for i in range(50):
        row = FIRST_ROW + i
        ws_eval.cell(row=row, column=1, value=i + 1)
        ws_eval.cell(row=row, column=2, value=user_msgs[i])
        ws_eval.cell(row=row, column=3, value=plain[i]["response"])
        ws_eval.cell(row=row, column=8, value=stateful[i]["response"])
        ws_eval.cell(row=row, column=13, value=sft[i]["response"])

    # Column widths
    ws_eval.column_dimensions["A"].width = 5
    ws_eval.column_dimensions["B"].width = 45
    for col_letter in ["C", "H", "M"]:
        ws_eval.column_dimensions[col_letter].width = 55
    for col in range(4, 8):
        ws_eval.column_dimensions[get_column_letter(col)].width = 10
    for col in range(9, 13):
        ws_eval.column_dimensions[get_column_letter(col)].width = 10
    for col in range(14, 18):
        ws_eval.column_dimensions[get_column_letter(col)].width = 10
    ws_eval.column_dimensions["R"].width = 30

    # Wrap text for response columns
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in range(FIRST_ROW, LAST_ROW + 1):
        for col in [2, 3, 8, 13]:
            ws_eval.cell(row=row, column=col).alignment = wrap

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Metric"
    ws_sum["B1"] = "Plain (A)"
    ws_sum["C1"] = "Stateful (B)"
    ws_sum["D1"] = "Stateful+SFT (C)"
    for cell in [ws_sum["A1"], ws_sum["B1"], ws_sum["C1"], ws_sum["D1"]]:
        cell.font = Font(bold=True)

    metrics = [
        ("Warmth",     "D", "I", "N"),
        ("Coherence",  "E", "J", "O"),
        ("Persona",    "F", "K", "P"),
        ("Engagement", "G", "L", "Q"),
    ]
    for row_idx, (name, col_a, col_b, col_c) in enumerate(metrics, start=2):
        ws_sum.cell(row=row_idx, column=1, value=name)
        ws_sum.cell(row=row_idx, column=2, value=f"=AVERAGE(Evaluation!{col_a}3:{col_a}52)")
        ws_sum.cell(row=row_idx, column=3, value=f"=AVERAGE(Evaluation!{col_b}3:{col_b}52)")
        ws_sum.cell(row=row_idx, column=4, value=f"=AVERAGE(Evaluation!{col_c}3:{col_c}52)")

    ws_sum.cell(row=6, column=1, value="Overall Average")
    ws_sum["A6"].font = Font(bold=True)
    ws_sum.cell(row=6, column=2, value="=AVERAGE(B2:B5)")
    ws_sum.cell(row=6, column=3, value="=AVERAGE(C2:C5)")
    ws_sum.cell(row=6, column=4, value="=AVERAGE(D2:D5)")

    wb.save(str(out_path))


def main() -> None:
    user_msgs = _load_user_messages()
    print(f"Loaded {len(user_msgs)} user messages\n")

    all_records: list[dict] = []

    # ── Plain ──
    print("=== Plain (7B base) ===")
    tokenizer, model = _load_model()
    plain = _generate_plain(tokenizer, model, user_msgs)
    all_records.extend(plain)
    _free_model(model, tokenizer)

    # ── Stateful ──
    print("\n=== Stateful (7B base) ===")
    tokenizer, model = _load_model()
    stateful = _generate_stateful(tokenizer, model, user_msgs, "stateful")
    all_records.extend(stateful)
    _free_model(model, tokenizer)

    # ── Stateful + SFT ──
    print("\n=== Stateful + SFT (7B LoRA) ===")
    tokenizer, model = _load_model(adapter_path=str(SFT_7B_PATH))
    sft = _generate_stateful(tokenizer, model, user_msgs, "stateful_sft")
    all_records.extend(sft)
    _free_model(model, tokenizer)

    # ── Save JSONL ──
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    jsonl_path = OUTPUT_DIR / "7b_human_eval_responses.jsonl"
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for rec in all_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"\nSaved {len(all_records)} records to {jsonl_path}")

    # ── Build xlsx ──
    xlsx_path = OUTPUT_DIR / "human_evaluation_7b.xlsx"
    _build_xlsx(user_msgs, plain, stateful, sft, xlsx_path)
    print(f"Saved evaluation sheet to {xlsx_path}")


if __name__ == "__main__":
    main()
